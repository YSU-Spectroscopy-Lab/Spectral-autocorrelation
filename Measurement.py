import random
import re
import shutil
import pandas as pd
import numpy as np
import xlwt
from keras.layers import Dense, Flatten, Conv1D, MaxPooling1D, Input
from keras import backend as K
from openpyxl import Workbook
from keras.models import Model
import os
from pandas import DataFrame


def read_single_component_data(path, newpath, txt_num):
    '''
    * read_Data.py file is placed under the same root directory as the read data folder.
    * path：Enter the read data folder path.
    * Read the data folder layout as shown in the example.
    * After rerunning to read the data, if you rerun to read, you need to delete the newly generated **_ok folder in the Data folder before starting the operation.
    '''
    os.mkdir(newpath)
    path = path
    file_name_lists = []
    for file_name in os.listdir(path):
        file_name_lists.append(file_name)
    for file_name_list in file_name_lists:
        root_ = []
        dirs_ = []
        a = 0
        for root, dirs, files in os.walk(path + '\\' + file_name_list):
            root_.append(root)
            dirs_.append(dirs)
            dirs__ = dirs_[0]
        file_name_after = newpath + '\\' + file_name_list + '_ok'
        if not os.path.exists(file_name_after):
            os.mkdir(file_name_after)
        for i in root_[1:]:
            file_name_after_ = file_name_after + '\\' + dirs__[a]
            if not os.path.exists(file_name_after_):
                os.mkdir(file_name_after_)
            txt_ = []
            for file_name_ in os.listdir(i):
                txt_.append(i + '\\' + file_name_)
            txt_num = txt_num  # 取多少文件
            txt_ = txt_[-txt_num:]
            num_all = []
            for txt_name in txt_:
                contents = []
                with open(txt_name, 'r') as f:
                    for line in f.readlines():
                        line = line.split('\n')
                        line = line[0].split('\t')
                        line = list(map(float, line))
                        contents.append(line)
                for content in contents:
                    num_all.append("%.4f" % (float(content[1]) / txt_num))
                if len(num_all) > len(contents):
                    for ii in range(len(num_all)):
                        if ii < len(contents):
                            num_all[ii] = "%.4f" % (num_all[ii] + float(num_all[ii + len(contents)]))
                        else:
                            num_all.pop()
                num_all = list(map(float, num_all))
                f.close()
                txt_name_after = newpath + '\\' + file_name_list + '_ok\\' + dirs__[a] + "\\" + dirs__[a] + ".txt"
                with open(txt_name_after, "w") as ff:
                    for li in num_all:
                        ff.write(str(li) + "\n")
                ff.close()
            a += 1
        print(file_name_list, "Data reading completed！")
    print("All  data read completed！")


def remove_bd(newpath):
    path = newpath
    nn_ = []
    file_name_lists = []
    for file_name in os.listdir(path):
        file_name_lists.append(file_name)
    for file_name_list in file_name_lists:
        n_p = path + '\\' + file_name_list + '_z'
        nn_.append(n_p)
        os.mkdir(n_p)
        root_ = []
        dirs_ = []

        for root, dirs, files in os.walk(path + '\\' + file_name_list):
            root_.append(root)
            dirs_.append(dirs)
            dirs__ = dirs_[0]
        root_.pop(0)
        root__ = root_[-1]
        dirs___ = dirs__[-1]
        root_.pop()
        dirs__.pop()
        bd_name = root__ + '\\' + dirs___ + '.txt'
        for i in range(len(root_)):
            data = []
            file_name = root_[i] + '\\' + dirs__[i] + '.txt'
            file_name_ = n_p + '\\' + dirs__[i] + '.txt'
            with open(bd_name) as bd_f:
                bd_file = bd_f.read().split('\n')
            bd_f.close()
            with open(file_name, 'r+') as f:
                file = f.read().split('\n')
            f.close()
            del (bd_file[-1])
            del (file[-1])
            bd_file = list(map(float, bd_file))
            file = list(map(float, file))
            for i in range(len(bd_file)):
                i_num = "%.4f" % ((file[i]) / (bd_file[i]))
                data.append(i_num)
            with open(file_name_, 'w') as f_:
                f_.truncate(0)
                for ii in data:
                    f_.write(ii + "\n")
            f_.close()
    print('Successfully removed the backing!')
    return nn_


def writeinexcel(path, nn):
    lu = []
    path = path
    le_ = 0
    le = 0
    wb1 = xlwt.Workbook(encoding='utf-8')
    w1 = wb1.add_sheet('one')
    ipath_ = path[0]
    file_name_lists_ = []
    file_name_lists = []
    for file_name_ in os.listdir(ipath_):
        file_name_ = re.sub('\D', '', file_name_)
        b = list(file_name_)
        # b.pop()
        ans = "".join(map(str, b))
        file_name_lists_.append(ans)
    file_name_lists_ = list(map(int, file_name_lists_))
    file_name_lists_.sort()
    le = len(file_name_lists_)
    for le_i in range(le):
        w1.write(0, le_i + le_, file_name_lists_[le_i])
    le_ = le_ + len(file_name_lists_)
    for file_name_ in os.listdir(ipath_):
        file_name_lists.append(file_name_)

    file_name_lists.sort(key=lambda x: int(x[:-nn]))
    for i_a in range(len(file_name_lists)):
        path_ = ipath_ + '\\' + file_name_lists[i_a]
        ii = 1
        for line in open(path_, encoding='utf-8'):
            if line == '\n':
                continue
            else:
                w1.write(ii, i_a, float(line))
                ii += 1

    wb1.save(path[0] + ".xls")
    p1 = path[0] + ".xls"
    lu.append(p1)

    if len(path) > 1:
        ipath__ = path[1]
        le__ = 0
        lei = 0
        wb2 = xlwt.Workbook(encoding='utf-8')
        w2 = wb2.add_sheet('one')
        file_name_lists_ = []
        file_name_lists = []
        for file_name_ in os.listdir(ipath__):
            file_name_ = re.sub('\D', '', file_name_)
            b = list(file_name_)
            # b.pop()
            ans = "".join(map(str, b))
            file_name_lists_.append(ans)
        file_name_lists_ = list(map(int, file_name_lists_))
        file_name_lists_.sort()
        lei = len(file_name_lists_)
        for le_i in range(lei):
            w2.write(0, le_i + le__, file_name_lists_[le_i])
        le__ = le__ + len(file_name_lists_)
        for file_name_ in os.listdir(ipath__):
            file_name_lists.append(file_name_)
        file_name_lists.sort(key=lambda x: int(x[:-nn]))
        for i_a in range(len(file_name_lists)):
            path_ = ipath__ + '\\' + file_name_lists[i_a]
            ii = 1
            for line in open(path_, encoding='utf-8'):
                if line == '\n':
                    continue
                else:
                    w2.write(ii, i_a, float(line))
                    ii += 1
                    # print(line)
        wb2.save(path[1] + ".xls")
        p2 = path[1] + ".xls"
        lu.append(p2)

    return lu


def koumanbian_no(x, data_all, a, b, e, f, i, j, m, n):
    aa = []
    j_ = 0

    line1 = list(range(a - 1, b))
    # print("line1:",len(line1))
    line2 = list(range(e - 1, f))
    # print("line2:", len(line2))
    line3 = list(range(i - 1, j))
    # print("line3:", len(line3))
    line4 = list(range(m - 1, n))
    # print("line4:", len(line4))
    line_z = list(range(a - 1, n))
    # print(len(line_z))

    listall = line1 + line2 + line3 + line4
    # print(len(listall))

    for i in x:
        zz1 = np.polyfit(listall, i, 3)  # 拟合 3
        pp1 = np.poly1d(zz1)
        jj = data_all[j_]

        # aa.append(i/pp1([i for i in range(len(i))]))
        aa.append(np.log(jj / pp1(line_z)))
        # aa.append(jj / pp1(line_z))
        j_ += 1

    return np.array(aa)


def dif_no(path):
    # Weak characteristic absorption region 1
    a = 512  # 512(193.77nm)
    b = 700

    # Characteristic absorption peak 1
    c = 701
    d = 737

    # Weak characteristic absorption region 2
    e = 738
    f = 888

    # Characteristic absorption peak 2
    g = 889
    h = 926

    # Weak characteristic absorption region 3
    i = 927
    j = 1106

    # Characteristic absorption peak 3
    k = 1107
    l = 1145

    # Weak characteristic absorption region 4
    m = 1146
    n = 1196

    data = pd.read_excel(path)
    columns = data.columns
    wb_all = Workbook()
    wb_all.create_sheet(index=0, title="all")
    ws_all = wb_all.active

    data_line_1 = data.iloc[a - 1:b, :]
    data_peak_1 = data.iloc[c - 1:d, :]
    data_line_2 = data.iloc[e - 1:f, :]
    data_peak_2 = data.iloc[g - 1:h, :]
    data_line_3 = data.iloc[i - 1:j, :]
    data_peak_3 = data.iloc[k - 1:l, :]
    data_line_4 = data.iloc[m - 1:n, :]

    data_line_1 = np.array(data_line_1)
    data_peak_1 = np.array(data_peak_1)
    data_line_2 = np.array(data_line_2)
    data_peak_2 = np.array(data_peak_2)
    data_line_3 = np.array(data_line_3)
    data_peak_3 = np.array(data_peak_3)
    data_line_4 = np.array(data_line_4)

    line_1_shape = data_line_1.shape[0]
    line_2_shape = data_line_2.shape[0]
    line_3_shape = data_line_3.shape[0]
    line_4_shape = data_line_4.shape[0]
    peak_1_shape = data_peak_1.shape[0]
    peak_2_shape = data_peak_2.shape[0]
    peak_3_shape = data_peak_3.shape[0]

    data_all = np.concatenate(
        (data_line_1, data_peak_1, data_line_2, data_peak_2, data_line_3, data_peak_3, data_line_4), axis=0)
    # print(data_all.shape)
    data_all = pd.DataFrame(data_all, columns=columns)
    data_all = data_all.T
    data_all = np.array(data_all)
    # print(data_all.shape)
    data_line_all = np.concatenate((data_line_1, data_line_2, data_line_3, data_line_4), axis=0)
    data_line_all = pd.DataFrame(data_line_all, columns=columns)
    data_line_all = data_line_all.T
    data_line_all = np.array(data_line_all)
    # print(data_line_all.shape)
    # print(data_all.shape)
    data_deal_all_line = koumanbian_no(data_line_all, data_all, a, b, e, f, i, j, m, n)
    deal_data_line_all = pd.DataFrame(data_deal_all_line)
    # print("deal_data_line_all",deal_data_line_all)
    columns_ = deal_data_line_all.columns
    # all_data
    l = len(columns_)
    i1 = 0
    for i in range(l):
        i1 += 1
        lie = deal_data_line_all[columns_[i]]
        for j in range(len(lie)):
            ws_all.cell(1, i + 1, i1)
            ws_all.cell(j + 2, i + 1, lie[j])
    for col_i in range(len(columns)):
        col = columns[col_i]
        ws_all.cell(col_i + 2, l + 1, col)
    ws_all.cell(1, l + 1, l + 1)
    name = path + '.xlsx'
    wb_all.save(name)
    name1 = to_pkl(name)
    print("NO Data differential completion!")
    return name1


def dif_nh3(path):
    # Weak characteristic absorption region 1
    a = 512  # 512(193.77nm)
    b = 700

    # Characteristic absorption peak 1
    c = 701
    d = 737

    # Weak characteristic absorption region 2
    e = 738
    f = 888

    # Characteristic absorption peak 2
    g = 889
    h = 926

    # Weak characteristic absorption region 3
    i = 927
    j = 1106

    # Characteristic absorption peak 3
    k = 1107
    l = 1145

    # Weak characteristic absorption region 4
    m = 1146
    n = 1196

    data = pd.read_excel(path)
    columns = data.columns
    wb_all = Workbook()
    wb_all.create_sheet(index=0, title="all")
    ws_all = wb_all.active

    data_line_1 = data.iloc[a - 1:b, :]
    data_peak_1 = data.iloc[c - 1:d, :]
    data_line_2 = data.iloc[e - 1:f, :]
    data_peak_2 = data.iloc[g - 1:h, :]
    data_line_3 = data.iloc[i - 1:j, :]
    data_peak_3 = data.iloc[k - 1:l, :]
    data_line_4 = data.iloc[m - 1:n, :]

    data_line_1 = np.array(data_line_1)
    data_peak_1 = np.array(data_peak_1)
    data_line_2 = np.array(data_line_2)
    data_peak_2 = np.array(data_peak_2)
    data_line_3 = np.array(data_line_3)
    data_peak_3 = np.array(data_peak_3)
    data_line_4 = np.array(data_line_4)

    data_all = np.concatenate(
        (data_line_1, data_peak_1, data_line_2, data_peak_2, data_line_3, data_peak_3, data_line_4), axis=0)
    # print(data_all.shape)
    data_all = pd.DataFrame(data_all, columns=columns)
    data_all = data_all.T
    data_all = np.array(data_all)
    # print(data_all.shape)
    data_line_all = np.concatenate((data_line_1, data_line_2, data_line_3, data_line_4), axis=0)
    data_line_all = pd.DataFrame(data_line_all, columns=columns)
    data_line_all = data_line_all.T
    data_line_all = np.array(data_line_all)
    # print(data_line_all.shape)
    # print(data_all.shape)
    data_deal_all_line = koumanbian_nh3(data_line_all, data_all, a, b, e, f, i, j, m, n)
    deal_data_line_all = pd.DataFrame(data_deal_all_line)
    # print("deal_data_line_all",deal_data_line_all)
    columns_ = deal_data_line_all.columns
    # all_data
    l = len(columns_)
    i1 = 0
    for i in range(l):
        i1 += 1
        lie = deal_data_line_all[columns_[i]]
        for j in range(len(lie)):
            ws_all.cell(1, i + 1, i1)
            ws_all.cell(j + 2, i + 1, lie[j])
    for col_i in range(len(columns)):
        col = columns[col_i]
        ws_all.cell(col_i + 2, l + 1, col)
    ws_all.cell(1, l + 1, l + 1)
    name = path + '.xlsx'
    wb_all.save(name)
    print("NH3 Data differential completion!")
    name1 = to_pkl(name)
    return name1


def koumanbian_nh3(x, data_all, a, b, e, f, i, j, m, n):
    aa = []
    j_ = 0

    line1 = list(range(a - 1, b))
    # print("line1:",len(line1))
    line2 = list(range(e - 1, f))
    # print("line2:", len(line2))
    line3 = list(range(i - 1, j))
    # print("line3:", len(line3))
    line4 = list(range(m - 1, n))
    # print("line4:", len(line4))
    line_z = list(range(a - 1, n))
    # print(len(line_z))

    listall = line1 + line2 + line3 + line4
    # print(len(listall))

    for i in x:
        zz1 = np.polyfit(listall, i, 3)  # 拟合 3
        pp1 = np.poly1d(zz1)
        jj = data_all[j_]

        # aa.append(i/pp1([i for i in range(len(i))]))
        aa.append(np.log(jj / pp1(line_z)))
        # aa.append(jj / pp1(line_z))
        j_ += 1
    return np.array(aa)


def to_pkl(path):
    # Read excel files
    df1 = DataFrame(pd.read_excel(path))
    dir_name = os.path.dirname(path)
    base_name = os.path.basename(path)
    suffix = base_name.split(".")[0]
    path_ = dir_name + "/" + suffix + ".pkl"
    df1.to_pickle(path_)
    return path_


def writeinexcel_multi(path3, path4, no_nh3_spectrum_path, add_data=True):
    wb = Workbook()
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    data1 = pd.read_excel(path3)
    data2 = pd.read_excel(path4)
    columns = data1.columns
    index1 = data1.index

    le = 0
    for i in range(data1.shape[1]):
        ws.cell(1, i + 1, columns[i])
    ws.cell(1, data1.shape[1] + 1, data1.shape[1] + 1)
    for i in range(data1.shape[0]):
        ind1 = list(data1.loc[index1[i]])
        # ind1.pop(-1)
        ind2 = list(data2.loc[index1[i]])
        # ind2.pop(-1)
        ind1 = np.array(ind1)
        ind2 = np.array(ind2)
        data_all = ind1 + ind2
        for j in range(len(data_all) - 1):
            ws.cell(i + 2, j + 1, data_all[j])
        ws.cell(i + 2, len(data_all), ind1[len(data_all) - 1])
        ws.cell(i + 2, len(data_all) + 1, ind2[len(data_all) - 1])

    le += int(data1.shape[0])
    for time in range(40):
        a = random.uniform(0, 1)
        a = round(a, 3)
        b = random.uniform(0, 1)
        b = round(b, 3)
        if a == 0 or b == 0:
            a += 0.1
            b += 0.1
        for i in range(data1.shape[0]):
            for j in range(data1.shape[0]):
                ind1 = list(data1.loc[index1[i]])
                ind2 = list(data2.loc[index1[j]])
                ind1 = np.array(ind1)
                ind2 = np.array(ind2)
                '''Combination Rules'''
                a = float(a)
                b = float(b)
                ind1 = a * ind1
                ind2 = b * ind2
                data_all = ind1 + ind2
                for iii in range(len(data_all) - 1):
                    ws.cell(le + 2 + j, iii + 1, data_all[iii])
                ws.cell(le + j + 2, len(data_all), ind1[len(data_all) - 1])
                ws.cell(le + j + 2, len(data_all) + 1, ind2[len(data_all) - 1])
            le += int(data1.shape[0])
        print("circulate", time, "time")
    wb.save(no_nh3_spectrum_path)
    to_pkl(no_nh3_spectrum_path)
    print("multicomponent Data expansion completed!")


def test_NO_model_cnn_single(path, model_stru1, model_para1, test_path, optimizer, loss, num):
    # Loading data
    df = pd.read_pickle(path)
    x_test = np.expand_dims(df.values[:, 0:-1].astype(float), axis=2)  # Adding a one-dimensional axis
    y_test = df.values[:, -1] / num
    # Divide training set, test set
    print("Loading of data complete!")
    x_test = x_test
    y_test = y_test
    loaded_model = load_first_cnn(model_stru1, model_para1)
    loaded_model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    predicted = predict_data(loaded_model, x_test, optimizer, loss)
    mean_error(predicted, y_test, num)
    save_model_excel(predicted, y_test, test_path, num)
    print("Test complete!")


def test_NH3_model_cnn_single(path, model_stru1, model_para1, test_path, optimizer, loss, num):
    # Loading data
    df = pd.read_pickle(path)
    x_test = np.expand_dims(df.values[:, 0:-1].astype(float), axis=2)  # Adding a one-dimensional axis
    y_test = df.values[:, -1] / num
    # Divide training set, test set
    print("Loading of data complete!")
    x_test = x_test
    y_test = y_test
    loaded_model = load_first_cnn(model_stru1, model_para1)
    loaded_model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    predicted = predict_data(loaded_model, x_test, optimizer, loss)
    mean_error(predicted, y_test, num)
    save_model_excel(predicted, y_test, test_path, num)
    print("Test complete!")


def load_first_cnn(model_stru2, model_para1):
    loaded_model = test_build_CNN_model()
    loaded_model.load_weights(model_para1)
    print("Loading of pre-trained models complete!")
    return loaded_model


def test_build_CNN_model():
    input1 = Input(shape=(685, 1))
    conv_layer1_1 = Conv1D(16, 3, strides=1, activation='relu')(input1)
    # conv_layer1_1_ = Conv1D(16, 3, strides=1, activation='relu')(conv_layer1_1)
    max_layer1_1 = MaxPooling1D(3)(conv_layer1_1)
    conv_layer1_2 = Conv1D(32, 3, strides=1, activation='relu')(max_layer1_1)
    # conv_layer1_2_ = Conv1D(32, 3, strides=1, activation='relu')(conv_layer1_2)
    max_layer1_2 = MaxPooling1D(3)(conv_layer1_2)
    conv_layer1_3 = Conv1D(32, 3, activation='relu')(max_layer1_2)
    # conv_layer1_3_ = Conv1D(32, 3, activation='relu')(conv_layer1_3)
    max_layer1_3 = MaxPooling1D(3)(conv_layer1_3)
    conv_layer1_4 = Conv1D(32, 3, activation='relu')(max_layer1_3)
    # conv_layer1_4_ = Conv1D(32, 3, activation='relu')(conv_layer1_4)
    max_layer1_4 = MaxPooling1D(3)(conv_layer1_4)
    flatten = Flatten()(max_layer1_4)
    f1 = Dense(1, activation='linear', name='prediction_one')(flatten)
    model = Model(outputs=f1, inputs=input1)
    model.summary()
    return model


# Custom metric function, determination factor R_Squares
def coeff_determination(y_true, y_pred):
    SS_res = K.sum(K.square(y_true - y_pred))
    SS_tot = K.sum(K.square(y_true - K.mean(y_true)))
    return (1 - SS_res / (SS_tot + K.epsilon()))


# Calculating the mean absolute error
def mean_error(predicted, y_test, num):
    y_test_size = np.reshape(y_test, (len(y_test), 1))
    result = np.mean(abs(predicted * num - y_test_size * num))
    print("MAE:", result)
    return result


# Predicted data
def predict_data(model, X_test, optimizer, loss):
    model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    predicted = model.predict(X_test)
    return predicted


# Save prediction results
def save_model_excel(predicted, Y_test, name, num):
    wb = Workbook()  # Create a new Excel file
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    Y_test = Y_test
    ws.cell(1, 1, 'Predicted concentration')
    ws.cell(1, 2, 'Real concentration')
    for i in range(len(predicted)):
        ws.cell(i + 2, 1, predicted[i][0] * num)
        ws.cell(i + 2, 2, Y_test[i] * num)
    wb.save(name)
    print("Save value to finish!")


def del_files(path):
    if os.path.exists(path):
        shutil.rmtree(path, ignore_errors=False, onerror=None)
    print("Data environment cleanup succeeded!")


def del_files_single(path1, path2):
    if os.path.exists(path1):
        shutil.rmtree(path1, ignore_errors=False, onerror=None)
    if os.path.exists(path2):
        shutil.rmtree(path2, ignore_errors=False, onerror=None)
    print("Data environment cleanup succeeded!")


if __name__ == '__main__':

    test_no_model = False
    test_nh3_model = False

    optimizer = "adam"
    loss = "mean_squared_error"
    Raw_data_path = "Measurement accuracy"
    Processed_data_path = "Measurement_accuracy_Processed_data"
    To_pkl_single_component_data_path = "Measurement_accuracy_single_component_data"
    Test_no_result_data_path = "Measurement_accuracy_NO_results_data"
    Test_nh3_result_data_path = "Measurement_accuracy_NH3_results_data"
    txt_num = 40
    numb = 9

    if test_no_model:
        del_files_single(Processed_data_path, To_pkl_single_component_data_path)
        read_single_component_data(Raw_data_path, Processed_data_path, txt_num)
        path1 = remove_bd(Processed_data_path)
        path2 = writeinexcel(path1, numb)
        path3 = dif_no(path2[1])

        del_files(Test_no_result_data_path)
        os.mkdir(Test_no_result_data_path)
        Pkl_data_path = path3
        result_values_file_name = Test_no_result_data_path + "/no-results.xlsx"
        model_para = "Result_no_cnn/no-para.h5"
        model_stru = "Result_no_cnn/no-stru.json"

        test_NO_model_cnn_single(Pkl_data_path, model_stru, model_para, result_values_file_name, optimizer, loss, 1000)

    if test_nh3_model:
        del_files_single(Processed_data_path, To_pkl_single_component_data_path)
        read_single_component_data(Raw_data_path, Processed_data_path, txt_num)
        path1 = remove_bd(Processed_data_path)
        path2 = writeinexcel(path1, numb)
        path4 = dif_nh3(path2[0])
        del_files(Test_nh3_result_data_path)
        os.mkdir(Test_nh3_result_data_path)
        Pkl_data_path = path4
        result_values_file_name = Test_nh3_result_data_path + "/nh3-results.xlsx"
        model_para = "Result_nh3_cnn/nh3-para.h5"
        model_stru = "Result_nh3_cnn/nh3-stru.json"
        test_NO_model_cnn_single(Pkl_data_path, model_stru, model_para, result_values_file_name, optimizer, loss, 10000)
