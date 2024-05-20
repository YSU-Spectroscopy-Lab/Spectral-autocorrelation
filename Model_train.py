import Model_build
import os
import shutil
from keras import backend as K
from openpyxl import Workbook
import pandas as pd
import numpy as np
from sklearn.metrics import r2_score
import matplotlib.pyplot as plt


def train_model(model, name, loss_name, model_para, model_stru, optimizer, loss, X_train, X_test, Y_train, Y_test,
                name1, name2, mae_r2_name):
    model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    history = model.fit(X_train, Y_train,
                        batch_size=128,
                        epochs=1000,
                        validation_data=(X_test, Y_test)
                        )
    predicted = predict_data(model, X_test, optimizer=optimizer, loss=loss)
    mae = mean_error(predicted, Y_test, num)
    save_model(model, model_stru, model_para)
    r2 = show_scores(predicted, Y_test)
    save_model_excel(predicted, Y_test, name, num)
    plot_history(history, name1, name2)
    save_model_loss(history, loss_name)
    save_mae_r2(mae, r2, mae_r2_name)


def train_model_multi(model, name, loss_name, model_para, model_stru, optimizer, loss, X_train, X_test, Y_train, Y_test,
                      name1, name2, mae_r2_name):
    model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    history = model.fit(X_train, Y_train,
                        batch_size=128,
                        epochs=200,
                        validation_data=(X_test, Y_test)
                        )
    predicted = predict_data(model, X_test, optimizer=optimizer, loss=loss)

    NO, NH3 = mean_error_multi(predicted, Y_test, num)
    no_r2, nh3_R2 = show_scores_multi(predicted, Y_test)
    save_model_excel_multi(predicted, Y_test, name, num)
    plot_history(history, name1, name2)
    save_model(model, model_stru, model_para)


# Custom metric function, determination factor R_Squares
def coeff_determination(y_true, y_pred):
    SS_res = K.sum(K.square(y_true - y_pred))
    SS_tot = K.sum(K.square(y_true - K.mean(y_true)))
    return (1 - SS_res / (SS_tot + K.epsilon()))


# Save prediction results
def save_model_excel(predicted, Y_test, name, num):
    wb = Workbook()  # Create a new Excel file
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    Y_test = Y_test
    ws.cell(1, 1, 'Predicted concentration')
    ws.cell(1, 2, 'Real concentration')
    for i in range(len(predicted)):
        ws.cell(i + 2, 1, predicted[i][0] * num)
        ws.cell(i + 2, 2, Y_test[i] * num)
    wb.save(name)
    print("Save value to finish!")


# Save prediction results
def save_model_excel_multi(predicted, Y_test, name, num):
    wb = Workbook()  # Create a new Excel file
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    Y_test = Y_test
    ws.cell(1, 1, 'NO Predicted concentration')
    ws.cell(1, 2, 'NH3 Predicted concentration')
    ws.cell(1, 3, 'NO Real concentration')
    ws.cell(1, 4, 'NH3 Real concentration')
    for i in range(predicted.shape[0]):
        ws.cell(i + 2, 1, predicted[i][0] * num)
        ws.cell(i + 2, 2, predicted[i][1] * num)
        ws.cell(i + 2, 3, Y_test[i][0] * num)
        ws.cell(i + 2, 4, Y_test[i][1] * num)
    wb.save(name)
    print("Save value to finish!")


def save_model_loss(history, name):
    wb = Workbook()  # Create a new Excel file
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    hist = pd.DataFrame(history.history)
    loss = hist['loss']
    val_loss = hist['val_loss']
    ws.cell(1, 1, 'loss')
    ws.cell(1, 2, 'val_loss')
    for i in range(len(loss)):
        ws.cell(i + 2, 1, loss[i])
        ws.cell(i + 2, 2, val_loss[i])
    wb.save(name)
    print("Save loss to finish!")


# Predicted data
def predict_data(model, X_test, optimizer, loss):
    model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    predicted = model.predict(X_test)
    return predicted


# Calculation of the decision factor
def show_scores(predicted, Y_test):
    r2_scores = r2_score(predicted, Y_test)
    print("R2:", r2_scores)
    return r2_scores


# Calculation of the decision factor
def show_scores_multi(predicted, Y_test):
    r2_scores_no = r2_score(predicted[:, 0], Y_test[:, 0])
    r2_scores_nh3 = r2_score(predicted[:, 1], Y_test[:, 1])
    print("NO R2:", r2_scores_no)
    print("NH3 R2:", r2_scores_nh3)
    return r2_scores_no, r2_scores_nh3


# Calculating the mean absolute error
def mean_error(predicted, y_test, num):
    y_test_size = np.reshape(y_test, (len(y_test), 1))
    result = np.mean(abs(predicted * num - y_test_size * num))
    print("MAE:", result)
    return result


# Calculating the mean absolute error
def mean_error_multi(predicted, y_test, num):
    result = np.mean(abs(predicted[:, 0] * num - y_test[:, 0] * num))
    print("NO MAE:", result)
    result_ = np.mean(abs(predicted[:, 1] * num - y_test[:, 1] * num))
    print("NH3 MAE:", result_)
    return result, result_


# Preservation of models
def save_model(model, name1, name2):
    # Convert their model grid structure to json storage
    # Store model parameter weights as h5 files
    model_json = model.to_json()
    with open(name1, 'w') as json_file:
        json_file.write(model_json)
    model.save_weights(name2)
    print("Save model complete!")


def plot_history(history, name1, name2):
    hist = pd.DataFrame(history.history)
    hist['epoch'] = history.epoch
    plt.figure()
    plt.xlabel('Epoch')
    plt.ylabel('loss')
    plt.plot(hist['epoch'], hist['loss'],
             label='Train loss')
    plt.plot(hist['epoch'], hist['val_loss'],
             label='Val loss')
    plt.ylim([-0.001, 0.05])
    plt.legend()
    plt.savefig(name1, dpi=600)
    plt.figure()
    plt.xlabel('Epoch')
    plt.ylabel('coeff_determination')
    plt.plot(hist['epoch'], hist['coeff_determination'],
             label='Train coeff_determination')
    plt.plot(hist['epoch'], hist['val_coeff_determination'],
             label='Val coeff_determination')
    plt.ylim([-0.5, 1.5])
    plt.legend()
    plt.savefig(name2, dpi=600)
    # plt.show()


def save_mae_r2(mae, r2, name):
    wb = Workbook()  # Create a new Excel file
    wb.create_sheet(index=0, title="all")
    ws = wb.active

    ws.cell(1, 1, 'MAE')
    ws.cell(1, 2, 'R2')
    ws.cell(2, 1, mae)
    ws.cell(2, 2, r2)
    wb.save(name)
    print("Save mae & r2 to finish!")


def del_files(path):
    if os.path.exists(path):
        shutil.rmtree(path, ignore_errors=False, onerror=None)
    print("Data environment cleanup succeeded!")


if __name__ == '__main__':

    Result_no_data_path = "Result_no_cnn"
    Result_nh3_data_path = "Result_nh3_cnn"
    Result_no_nh3_data_path = "Result_no_nh3_cnn"

    optimizer = "adam"
    loss = "mean_squared_error"
    train_no_model = False
    train_nh3_model = False
    train_no_nh3_model = False

    if train_no_model:
        del_files(Result_no_data_path)
        os.mkdir(Result_no_data_path)
        model_structure = Result_no_data_path + "/CNN.png"
        result_values_file_name = Result_no_data_path + "/no-results.xlsx"
        loss_file_name = Result_no_data_path + "/no-loss.xlsx"
        model_para = Result_no_data_path + "/no-para.h5"
        model_stru = Result_no_data_path + "/no-stru.json"
        model_loss = Result_no_data_path + "/no-loss.png"
        model_R2 = Result_no_data_path + "/no-R2.png"
        model_mae_r2_name = Result_no_data_path + "/no-mae-r2.xlsx"
        model, X_train, X_test, Y_train, Y_test, num = Model_build.run(train_no_model, train_nh3_model,
                                                                       train_no_nh3_model, model_structure)
        train_model(model, result_values_file_name, loss_file_name, model_para, model_stru, optimizer, loss, X_train,
                    X_test, Y_train,
                    Y_test, model_loss, model_R2, model_mae_r2_name)

    if train_nh3_model:
        del_files(Result_nh3_data_path)
        os.mkdir(Result_nh3_data_path)
        model_structure = Result_nh3_data_path + "/CNN.png"
        result_values_file_name = Result_nh3_data_path + "/nh3-results.xlsx"
        loss_file_name = Result_nh3_data_path + "/nh3-loss.xlsx"
        model_para = Result_nh3_data_path + "/nh3-para.h5"
        model_stru = Result_nh3_data_path + "/nh3-stru.json"
        model_loss = Result_nh3_data_path + "/nh3-loss.png"
        model_R2 = Result_nh3_data_path + "/nh3-R2.png"
        model_mae_r2_name = Result_nh3_data_path + "/nh3-mae-r2.xlsx"
        model, X_train, X_test, Y_train, Y_test, num = Model_build.run(train_no_model, train_nh3_model,
                                                                       train_no_nh3_model, model_structure)
        train_model(model, result_values_file_name, loss_file_name, model_para, model_stru, optimizer, loss, X_train,
                    X_test, Y_train,
                    Y_test, model_loss, model_R2, model_mae_r2_name)

    if train_no_nh3_model:
        del_files(Result_no_nh3_data_path)
        os.mkdir(Result_no_nh3_data_path)
        model_structure = Result_no_nh3_data_path + "/CNN.png"
        result_values_file_name = Result_no_nh3_data_path + "/no-h3-results.xlsx"
        loss_file_name = Result_no_nh3_data_path + "/no-nh3-loss.xlsx"
        model_para = Result_no_nh3_data_path + "/no-nh3-para.h5"
        model_stru = Result_no_nh3_data_path + "/no-nh3-stru.json"
        model_loss = Result_no_nh3_data_path + "/no-nh3-loss.png"
        model_R2 = Result_no_nh3_data_path + "/no-nh3-R2.png"
        model_mae_r2_name = Result_no_nh3_data_path + "/no-nh3-mae-r2.xlsx"
        model, X_train, X_test, Y_train, Y_test, num = Model_build.run(train_no_model, train_nh3_model,
                                                                       train_no_nh3_model, model_structure)
        train_model_multi(model, result_values_file_name, loss_file_name, model_para, model_stru, optimizer, loss,
                          X_train, X_test, Y_train,
                          Y_test, model_loss, model_R2, model_mae_r2_name)
